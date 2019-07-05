import re
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter

# 提取路名的信息，并截掉该字段
def extract_road(str_in):
    road = ' '
    road_pattern = re.compile("(.+?[路道])")  # xxx路|道， 取出第一个 路|道 及之前的内容
    road_pattern_1 = re.compile("(.*?[^\d][街巷弄])")  # xxx街|巷|弄,取出第一个 街|巷|弄 及之前的内容，且之前的内容不能有数字
    road_match = road_pattern.match(str_in)
    road_match_1 = road_pattern_1.match(str_in)
    if road_match:
        road = road_match.group(1)
    elif road_match_1:
        road = road_match_1.group(1)
    else:
        pass
    index = str_in.find(road)
    str_in = str_in[index + len(road):]
    return str_in,road

# 提取村名的信息，并截掉该字段
def extract_village(str_in):
    village = ' '
    #village_pattern = re.compile(".+?\d+(.+?[乡屯村]).*?")  # 若为 数字xx村，取出中间的xx村的内容
    village_pattern = re.compile("(.*?[村乡屯])") # xxx村|乡|屯 取出第一个 村|乡|屯 及之前的内容
    special_pattern = re.compile("(.*?[村乡屯])[路街]") # 特殊情况 新村路 等
    village_pattern_clear = re.compile(".*[号弄区路](.*)") # 去除村前面的多余信息
    village_match = village_pattern.match(str_in)
    special_match = special_pattern.match(str_in)
    if village_match:
        village = village_match.group(1)
    else:
        pass
    if special_match:
        village = ' '
    index = str_in.find(village)
    str_in = str_in[index + len(village):]
    # 去除 村 多余的信息
    village_match_clear = village_pattern_clear.match(village)
    if village_match_clear:
        village = village_match_clear.group(1)
    return str_in,village

# 提取路名和村名
def extract_village_road(str_in):
    roadFirst_pattern = re.compile(".+?[路街道巷].+?[村乡屯]")
    villageFirst_pattern = re.compile(".+?[村乡屯].+?[路街道巷]")
    village = ' '
    road = ' '
    roadFirst_match = roadFirst_pattern.match(str_in)
    villageFirst_match = villageFirst_pattern.match(str_in)
    # 如果路名在村名前面
    if roadFirst_match:
        str_in, road = extract_road(str_in)
        str_in, village = extract_village(str_in)
    # 如果村名在路名前面
    elif villageFirst_match:
        str_in, village = extract_village(str_in)
        str_in, road = extract_road(str_in)
    else:
        # 有 村|乡|屯 关键字,则先进行村的检索
        if '村' in str_in or '乡' in str_in or '屯' in str_in: #or '庄' in str_in or '里' in str_in:
            str_in, village = extract_village(str_in)
        # 进行路的检索
        str_in, road = extract_road(str_in)
    # 去除公交车号被误认为是路的（181路）
    road_pattern_clear = re.compile(".*?\d")
    road_match_clear = road_pattern_clear.match(road)
    if road_match_clear:
        road = ' '
    return village, road

# txt转xlsx
def txt_to_xlsx(filename, outfile):
    fr = open(filename, 'r+', encoding='utf8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('^')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)

if __name__ == "__main__":
    new_txt = '/Users/Shar/Desktop/Sheets/上海_公司_补充_标准.txt'
    new_sheet = '/Users/Shar/Desktop/Sheets/上海_公司_补充_标准.xlsx'
    std_file = open(new_txt, 'w', encoding='utf8')
    data = xlrd.open_workbook("/Users/Shar/Desktop/Sheets/上海公司地址补充.xlsx")
    #data = xlrd.open_workbook("/Users/Shar/Desktop/Sheets/Debug.xlsx")
    table = data.sheets()[0]
    nrows = table.nrows

    region_pattern = re.compile("(.+?(镇|街道))[^路]") # 根据关键字 镇|街道 进行检索和分隔，且后面不为路

    parenthesis_pattern = re.compile(".*?\((.*?)\)") # xxx(xx), 取出括号里的内容
    parenthesis_out_pattern = re.compile("(.*?)\(.*?\)") # 截取括号外的内容

    num_left_pattern = re.compile(".+?[路村道街巷镇](.+)") # 取出 路|村|道|街|巷 后面的内容
    num_pattern = re.compile(".*?(\d+)(号|弄|幢|临|-|$)") # 取出数字，且后面不跟‘米’
    first_ten_pattern = re.compile("(十)[^一二三四五六七八九]")
    first_ten_1_pattern = re.compile("(十)[一二三四五六七八九]")
    middle_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[一二三四五六七八九]")
    last_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[^一二三四五六七八九]")
    count = 0
    for i in range(nrows):
        count = count + 1
        temp = str(table.row_values(i)[4]).strip()
        city = str(table.row_values(i)[2]).strip()
        district = str(table.row_values(i)[3]).strip()
        region_name = ' '   # 镇|街道
        village_name = ' ' # 村|乡|屯
        road_name = ' ' # 路
        num_str = ' ' # 号
        # 如果地址信息不为空
        if not temp.isspace():

            # 去掉区市信息
            if district in temp:
                index = temp.rfind(district)
                temp = temp[index+len(district):]
            elif city in temp:
                index = temp.rfind(city)
                temp = temp[index+len(city):]

            # temp里的市、区内容已被去除
            # 将剩下的内容分为括号里和括号外
            temp_out_match = parenthesis_out_pattern.match(temp)
            if temp_out_match:
                temp_out = temp_out_match.group(1)
            else:
                temp_out = temp
            parenthesis_match = parenthesis_pattern.match(temp)
            if parenthesis_match:
                temp_in = parenthesis_match.group(1)
            else:
                temp_in = ' '

            # 先处理括号外的字段，取出 镇|街道 的信息，并去除这个字段
            region_match = region_pattern.match(temp_out)
            if region_match:
                region_name = region_match.group(1)
                index = temp_out.find(region_name)
                temp_out = temp_out[index + len(region_name):]

            # 对括号外的字段提取 村 和 路 的信息
            village_name, road_name = extract_village_road(temp_out)

            # 如果村名为空，看能不能括号里提取出村名
            if village_name == ' ' and temp_in != ' ':
                trash,village_name = extract_village(temp_in)
            # 如果路名为空，看能不能怂括号里提取出路名
            if road_name == ' 'and temp_in != ' ':
                trash,road_name = extract_road(temp_in)

            # if village_name == '新村' or village_name == '府村'or village_name == '村' or village_name == '镇':
            #     village_name = ' '

            # 从temp里取出号
            # 把 路|村|道|街|巷|弄 后面的内容取出来
            num_left = ' '
            num_left_match = num_left_pattern.match(temp)
            if num_left_match:
                num_left = num_left_match.group(1)
            # 取出来的内容进行文字->阿拉伯数字的转换
            if num_left != ' ':
                first_ten = first_ten_pattern.match(num_left)
                first_ten_1 = first_ten_1_pattern.match(num_left)
                middle_ten = middle_ten_pattern.match(num_left)
                last_ten = last_ten_pattern.match(num_left)

                if first_ten:
                    num_left = num_left.replace("十", "10")
                elif first_ten_1:
                    num_left = num_left.replace("十", "1")
                elif middle_ten:
                    num_left = num_left.replace("十", "")
                elif last_ten:
                    num_left = num_left.replace("十", "0")

                num_left = num_left.replace("一", "1")
                num_left = num_left.replace("二", "2")
                num_left = num_left.replace("三", "3")
                num_left = num_left.replace("四", "4")
                num_left = num_left.replace("五", "5")
                num_left = num_left.replace("六", "6")
                num_left = num_left.replace("七", "7")
                num_left = num_left.replace("八", "8")
                num_left = num_left.replace("九", "9")
                num_left = num_left.replace("〇", "0")

                # 转换好以后的字段，取出第一个数字前的所有内容
                match_num = num_pattern.match(num_left)
                if match_num:
                    num_str = match_num.group(1)
        else:
            pass
        print(str(count) + "   " + temp + "              " + region_name + "              " + village_name + "              " + road_name + "           " + num_str)
        std_file.write(str(table.row_values(i)[0]).strip() + '^' + str(table.row_values(i)[1]).strip() + '^' +
                        str(table.row_values(i)[2]).strip() + '^' + str(table.row_values(i)[3]).strip() + '^' +
                        str(table.row_values(i)[4]).strip() + '^' + str(table.row_values(i)[5]).strip() + '^' +
                        str(table.row_values(i)[6]).strip() + '^' + region_name + '^' + village_name + '^' + road_name + '^' + num_str + '\n')
    print("Successfully saved to " + new_txt)
    txt_to_xlsx(new_txt, new_sheet)
    print("Successfully converted to " + new_sheet)
    print("-------------------------------------ALL JOBS DONE-----------------------------------")
    std_file.close()